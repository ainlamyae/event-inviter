"""Generate the example event-inviter workbook.

Produces event-inviter-template.xlsx next to this script, with two tabs:
  - Locations: Label | Address | Map URL
  - Event: 22 columns matching EVENT_FIELDS in js/google-api.js (Date, Title,
    Outline, Notes, Background, Links, Receivers, Hijri/Shamsi Date, Day of
    Week, Fajr/Dhuhr/Maghrib Time, 7 "show in email" flags, Sender Name,
    Closing).

Run: python generate_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "event-inviter-template.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_TOP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)


def write_sheet(ws, headers, rows, col_widths, rtl_columns=()):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = WRAP_TOP_RIGHT if col_idx in rtl_columns else WRAP_TOP_LEFT

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


def build_workbook():
    wb = Workbook()

    locations_ws = wb.active
    locations_ws.title = "Locations"
    write_sheet(
        locations_ws,
        headers=["Label", "Address", "Google Map URL"],
        rows=[
            [
                "MC 2017",
                "Mathematics and Computer Building, University Ave W, Waterloo, ON N2L 3L3",
                "https://maps.app.goo.gl/VNh5R6kFM1NbFKyY9",
            ],
            [
                "MC 2018",
                "Mathematics and Computer Building, University Ave W, Waterloo, ON N2L 3L3",
                "https://maps.app.goo.gl/VNh5R6kFM1NbFKyY9",
            ],
            [
                "DC 1350",
                "Davis Centre, University Ave W, Waterloo, ON N2L 3G1",
                "https://maps.app.goo.gl/2yV8ZbQGqQqQqQqQ7",
            ],
        ],
        col_widths=[14, 55, 42],
    )

    event_ws = wb.create_sheet("Event")

    # Agenda lines use the "HH:MM - HH:MM - Topic (Location)" format the app's
    # structured Agenda table serializes to (Persian or Latin digits both work).
    outline_text = (
        "۱۹:۴۵ - ۲۰:۳۰ - قرائت قرآن و تفسیر آیات ۲۱ تا ۲۶ سوره توبه (MC 2018)\n"
        "۲۰:۳۰ - ۲۰:۵۰ - پذیرایی (MC 2018)\n"
        "۲۰:۵۰ - ۲۱:۳۰ - ادامه ی تفسیر (MC 2018)\n"
        "۲۱:۳۰ - ۲۱:۴۵ - نماز جماعت (MC 2018)"
    )
    notes_text = "لطفاً به موقع حضور یابید."  # one Recommendation item, one bullet in the email
    background_text = (
        "رَبِّ اجْعَلْنِی "
        "مُقیمَ الصَّلاةِ "
        "وَمِن ذُرِّيَتِی "
        "ـ رَبَّنَا وَتَقَبَّل "
        "دُعَاءِ ﴾ابراهیم-۴۰﴿\n"
        "پروردگارا، مرا برپادارنده "
        "نماز قرار ده، و از فرزندان "
        "من نیز. پروردگارا، و دعای "
        "مرا بپذیر."
    )
    links_text = "جدول جلسات هفتگی: <insert your schedule link here>"
    closing_text = "با تشکر از حضور شما، منتظر دیدارتان هستیم."
    sender_name = "جلسه قرآن دانشگاه واترلو (University of Waterloo Quran Session)"

    write_sheet(
        event_ws,
        headers=[
            "Date", "Title", "Outline", "Notes", "Background", "Links", "Receivers",
            "Hijri Date", "Shamsi Date", "Day of Week",
            "Fajr Time", "Dhuhr Time", "Maghrib Time",
            "Show Gregorian In Email", "Show Hijri In Email", "Show Shamsi In Email",
            "Show Day Of Week In Email", "Show Fajr In Email", "Show Dhuhr In Email",
            "Show Maghrib In Email",
            "Sender Name", "Closing",
        ],
        rows=[
            [
                "2026-07-03",
                "جلسه هفتگی قرآن",
                outline_text,
                notes_text,
                background_text,
                links_text,
                "uw_q...@googlegroups.com",
                "۱۸ محرم ۱۴۴۸",
                "۱۲ تیر ۱۴۰۵",
                "جمعه (Friday)",
                "۰۳:۳۱",
                "۱۳:۲۷",
                "۲۱:۳۱",
                True, False, False, False, False, False, False,
                sender_name,
                closing_text,
            ],
        ],
        col_widths=[
            12, 22, 46, 26, 34, 30, 24,
            14, 14, 16,
            10, 10, 12,
            10, 10, 10, 10, 10, 10, 10,
            30, 30,
        ],
        rtl_columns=(2, 3, 4, 5, 6, 8, 9, 10, 21, 22),
    )

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
